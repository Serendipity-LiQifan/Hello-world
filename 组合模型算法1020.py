# noinspection PyUnresolvedReferences
import openpyxl

wb1 =openpyxl.load_workbook('example1020.xlsx')
sheet = wb1.get_sheet_by_name('汇总')
sheet1 =wb1.get_sheet_by_name('电流诊断')
sheet2 =wb1.get_sheet_by_name('多参诊断')
sheet3 =wb1.get_sheet_by_name('振动诊断')
sheet4 =wb1.get_sheet_by_name('憋压诊断')
#sheet5 =wb1.get_sheet_by_name('耗电量诊断')
Maxrow =(sheet1.max_row)-1
#Maxcolume1 =(sheet1.max_column)
#Maxcolume2 =(sheet2.max_column)
#Maxcolume3 =(sheet3.max_column)
#Maxcolume4 =(sheet4.max_column)
#Maxcolume5 =(sheet5.max_column)
#print(Maxrow)

# noinspection PyUnresolvedReferences
import xlrd
#for i in range(1,Maxrow):

book = xlrd.open_workbook("example1020.xlsx")

sheet11 = book.sheet_by_index(1)
sheet12 = book.sheet_by_index(2)
sheet13 = book.sheet_by_index(3)
sheet14 = book.sheet_by_index(4)
#sheet15 = book.sheet_by_index(5)
#读取表格数据：各诊断方法概率最大故障
M1 = sheet11.col_values(colx=11,start_rowx=1)
M2 = sheet12.col_values(colx=15,start_rowx=1)
M3 = sheet13.col_values(colx=5,start_rowx=1)
M4 = sheet14.col_values(colx=9,start_rowx=1)
#M5 = sheet15.col_values(colx=4,start_rowx=1)
print(f"电流（M1）诊断结果: {(M1)}")
print(f"多参（M2）诊断结果: {(M2)}")
print(f"振动（M3）诊断结果: {(M3)}")
print(f"憋压（M4）诊断结果: {(M4)}")
#print(f"耗电量（M5）诊断结果: {(M5)}")

#计算每个诊断结果的误判率
M11 = [1-i for i in M1]
M12 = [round(i,4) for i in M11]
M21 = [1-i for i in M2]
M22 = [round(i,4) for i in M21]
M31 = [1-i for i in M3]
M32 = [round(i,4) for i in M31]
M41 = [1-i for i in M4]
M42 = [round(i,4) for i in M41]
#M51 = [1-i for i in M5]
#M52 = [round(i,4) for i in M51]

print(M12)
print(M22)
print(M32)
print(M42)
#print(M52)
#建立矩阵
# noinspection PyUnresolvedReferences
import numpy as np
e1 =np.array([M12,M22,M32,M42])
#输出误判率矩阵
e =e1.T
print(f"e矩阵为: {(e)}")
#建立误差信息矩阵E
E =e1.dot(e)
print(f"E矩阵为:{(E)}")
#创建列向量R
R =np.mat([[1],[1],[1],[1]])
Rz=R.T
#E矩阵的逆
En =np.linalg.inv(E)
#最优权值W的计算
W1 =En.dot(R)
W2 =Rz.dot(W1)
W =W1/W2
print(f"最优权值W矩阵为:{(W)}")

#输入数据
N1 =input('请输入数据的电流诊断结果：')
N2 =input('请输入数据的多参诊断结果：')
N3 =input('请输入数据的振动诊断结果：')
N4 =input('请输入数据的憋压诊断结果：')
result1 = {'正常': 1,
           '卡泵停机': 2,
           '过载停机': 3,
           '供液不足': 4,
           '气体影响': 5,
           '频繁起停': 6,
           '出砂': 7,
           '气锁': 8,
           '负载增加': 9,
           '电压波动': 10}
result2 = {'正常': 11,
           '含水变化': 12,
           '地面控制异常': 13,
           '供液不足': 14,
           '叶轮磨损': 15,
           '泵反转': 16,
           '出砂': 17,
           '气体影响': 18,
           '轴断脱': 19,
           '泵入口堵': 20,
           '稠油及乳化': 21,
           '气锁': 22,
           '泵内堵塞': 23,
           '管柱漏失': 24}
result3 ={'正常': 25,
          '泵内堵塞': 26,
          '叶轮磨损': 27,
          '轴断脱': 28}
result4 ={'正常': 29,
          '管柱漏失': 30,
          '叶轮磨损': 31,
          '轴断脱': 32,
          '结垢': 33,
          '稠油及乳化': 34,
          'Y接头漏失': 35,
          '气体影响': 36}
N11 = result1[N1]
N12 = result2[N2]
N13 = result3[N3]
N14 = result4[N4]
print(N11)
print(N12)
print(N13)
print(N14)



#N1 =int(N)
#x1= sheet11.row_values(rowx=N1)
#x2= sheet12.row_values(rowx=N1)
#x3= sheet13.row_values(rowx=N1)
#x4= sheet14.row_values(rowx=N1)
#x5= sheet15.row_values(rowx=N1)
#X1 =x1[1:10]
#print(X1)
#y1 = X1.index(max(X1))+1
#X2 =x2[1:5]
#print(X2)
#y2 = X2.index(max(X2))+10
#X3 =x3[1:4]
#print(X3)
#y3 = X3.index(max(X3))+14
#X4 =x4[1:2]
#print(X4)
#y4 = X4.index(max(X4))+17
#X5 =x5[1:3]
#print(X5)
#y5 = X5.index(max(X5))+18
#print(y1+1)
#print(y2+10)
#print(y3+12)
#print(y4+15)
#print(y5+16)
Y =[N11,N12,N13,N14]
print(f"将故障用数字表达：{(Y)}")

#构建数据矩阵
X01 =[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
X02 =[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
X03 =[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
X04 =[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
#X05 =[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
X0601 =Y[0]
X0602 =Y[1]
X0603 =Y[2]
X0604 =Y[3]
#X0605 =Y[4]
print(X0601)
print(X0602)
print(X0603)
print(X0604)
#print(X0605)
X01[X0601-1] =1
X02[X0602-1] =1
X03[X0603-1] =1
X04[X0604-1] =1
#X05[X0605-1] =1
#输出专家诊断矩阵
X =np.array([X01,X02,X03,X04])
print(f"X矩阵为:{(X)}")
W01 =W.T#W转置
print(W01)
Z =W01.dot(X)
print(Z)
Z1 =Z.tolist()
print(Z1)
Z2 =Z1[0]
print(Z2)

#输出诊断结果
import heapq
result =map(Z2.index, heapq.nlargest(1,Z2))
RESULT0 =set(result)
RESULT00 =list(RESULT0)
RESULT01 =[str(i)for i in RESULT00]
print(RESULT01)

RESULT1 =['正常' if i =='0' else i for i in RESULT01]
RESULT2 =['卡泵停机' if i =='1' else i for i in RESULT1]
RESULT3 =['过载停机' if i =='2' else i for i in RESULT2]
RESULT4 =['供液不足（电流）' if i =='3' else i for i in RESULT3]
RESULT5 =['气体影响（电流）' if i =='4' else i for i in RESULT4]
RESULT6 =['频繁起停' if i =='5' else i for i in RESULT5]
RESULT7 =['出砂（电流）' if i =='6' else i for i in RESULT6]
RESULT8 =['气锁（电流）' if i =='7' else i for i in RESULT7]
RESULT9 =['负载增加' if i =='8' else i for i in RESULT8]
RESULT10 =['电压波动' if i =='9' else i for i in RESULT9]
RESULT11 =['正常' if i =='10' else i for i in RESULT10]
RESULT12 =['含水变化' if i =='11' else i for i in RESULT11]
RESULT13 =['地面控制异常' if i =='12' else i for i in RESULT12]
RESULT14 =['供液不足（多参）' if i =='13' else i for i in RESULT13]
RESULT15 =['叶轮磨损（多参）' if i =='14' else i for i in RESULT14]
RESULT16 =['泵反转' if i =='15' else i for i in RESULT15]
RESULT17 =['出砂（多参）' if i =='16' else i for i in RESULT16]
RESULT18 =['气体影响（多参）' if i =='17' else i for i in RESULT17]
RESULT19 =['轴断脱（多参）' if i =='18' else i for i in RESULT18]
RESULT20 =['泵入口堵' if i =='19' else i for i in RESULT19]
RESULT21 =['稠油及乳化（多参）' if i =='20' else i for i in RESULT20]
RESULT22 =['气锁（多参）' if i =='21' else i for i in RESULT21]
RESULT23 =['泵内堵塞（多参）' if i =='22' else i for i in RESULT22]
RESULT24 =['管柱漏失（多参）' if i =='23' else i for i in RESULT23]
RESULT25 =['正常' if i =='24' else i for i in RESULT24]
RESULT26 =['泵内堵塞（振动）' if i =='25' else i for i in RESULT25]
RESULT27 =['叶轮磨损（振动）' if i =='26' else i for i in RESULT26]
RESULT28 =['轴断脱（振动）' if i =='27' else i for i in RESULT27]
RESULT29 =['正常' if i =='28' else i for i in RESULT28]
RESULT30 =['管柱漏失（憋压）' if i =='29' else i for i in RESULT29]
RESULT31 =['叶轮磨损（憋压）' if i =='30' else i for i in RESULT30]
RESULT32 =['轴断脱（憋压）' if i =='31' else i for i in RESULT31]
RESULT33 =['结垢' if i =='32' else i for i in RESULT32]
RESULT34 =['稠油及乳化（憋压）' if i =='33' else i for i in RESULT33]
RESULT35 =['Y接头漏失' if i =='34' else i for i in RESULT34]
RESULT36 =['气体影响（憋压）' if i =='35' else i for i in RESULT35]
print(RESULT36)










